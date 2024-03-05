import openpyxl

# file_path = "C:\Selenium_Practice_Data_Files\TestExcelFile.xlsx"
# workbook = openpyxl.load_workbook(file_path)
# sheet = workbook.active   #---(or) sheet = workbook['Data'] #To get active sheet from excel

#Writing data
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value ="welcome"
# workbook.save(file_path)

#Multiple Data Adding
file_path = "C:\Selenium_Practice_Data_Files\TestExcelFile1.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active   #---(or) sheet = workbook['Data'] #To get active sheet from excel

sheet.cell(1,1).value = "100"
sheet.cell(1,2).value = "Amrutha"
sheet.cell(1,3).value = "QA"

sheet.cell(2,1).value = "101"
sheet.cell(2,2).value = "Asma"
sheet.cell(2,3).value = "Developer"

sheet.cell(3,1).value = "103"
sheet.cell(3,2).value = "Nazma"
sheet.cell(3,3).value = "Automation Tester"

workbook.save(file_path)  #Save the file after entering the data
