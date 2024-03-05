import openpyxl

#File --> Workbook ---> Sheets ---Rows --Cells
file_path = "C:\Selenium_Practice_Data_Files\Data.xlsx"

workbook = openpyxl.load_workbook(file_path)
sheet = workbook["Sheet1"]

rows = sheet.max_row
columns = sheet.max_column

#Reading all the rows & columns from the excel

for i in range(1, rows+1):
    for j in range(1,columns+1):
        print(sheet.cell(i,j).value, end="             " )
    print()
